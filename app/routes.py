from flask import Blueprint, render_template, redirect, url_for, request, flash, send_file
from flask_login import login_user, logout_user, login_required
from werkzeug.utils import secure_filename
from sqlalchemy import func, case
import openpyxl
import os
import io
from .models import User, StockItem
from . import db, login_manager

main = Blueprint('main', __name__)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@main.route('/')
def home():
    return render_template('home.html')

@main.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and user.password == password:
            login_user(user)
            return redirect(url_for('main.dashboard'))
        else:
            flash('Invalid credentials')
    return render_template('login.html')

@main.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

@main.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('main.home'))

@main.route('/upload', methods=['GET', 'POST'])
@login_required
def upload():
    if request.method == 'POST':
        file = request.files['file']
        if file and file.filename.endswith('.xlsx'):
            filename = secure_filename(file.filename)
            filepath = os.path.join('uploads', filename)
            os.makedirs('uploads', exist_ok=True)
            file.save(filepath)
            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                item = StockItem(
                    item_name=row[0],
                    company_name=row[1],
                    model=row[2],
                    client_name=row[3],
                    project_name=row[4],
                    quantity=int(row[5]) if row[5] else 0,
                    direction=row[6] if len(row) > 6 else 'IN'
                )
                db.session.add(item)
            db.session.commit()
            flash('Data uploaded successfully!')
            return redirect(url_for('main.dashboard'))
        else:
            flash('Please upload a valid .xlsx file')
    return render_template('upload.html')

@main.route('/stock', methods=['GET', 'POST'])
@login_required
def stock():
    if request.method == 'POST':
        identifier = request.form['item_identifier'].strip()
        action = request.form['action']
        qty = int(request.form['quantity'])

        item = StockItem.query.filter(
            (StockItem.item_name == identifier) | (StockItem.barcode == identifier)
        ).order_by(StockItem.timestamp.desc()).first()

        if not item:
            flash("Item not found. Please upload it first or check spelling.")
            return redirect(url_for('main.stock'))

        if action == 'in':
            item.quantity += qty
        elif action == 'out':
            if item.quantity < qty:
                flash("Not enough stock to remove!")
                return redirect(url_for('main.stock'))
            item.quantity -= qty

        movement = StockItem(
            item_name=item.item_name,
            barcode=item.barcode,
            company_name=item.company_name,
            client_name=item.client_name,
            project_name=item.project_name,
            collected_by=request.form['collected_by'],
            quantity=item.quantity
        )
        db.session.add(movement)
        db.session.commit()
        flash("Stock updated successfully!")
    return render_template('stock.html')

@main.route('/stock_report', methods=['GET'])
@login_required
def stock_report():
    query = StockItem.query
    search = request.args.get('search')
    if search:
        query = query.filter(
            (StockItem.item_name.ilike(f"%{search}%")) |
            (StockItem.client_name.ilike(f"%{search}%")) |
            (StockItem.project_name.ilike(f"%{search}%"))
        )
    all_items = query.order_by(StockItem.timestamp.desc()).all()
    latest_items = {}
    for item in all_items:
        if item.item_name not in latest_items:
            latest_items[item.item_name] = item
    return render_template('stock_report.html', items=latest_items.values(), search=search)

@main.route('/add_stock', methods=['GET', 'POST'])
@login_required
def add_stock():
    if request.method == 'POST':
        item = StockItem(
            client_name=request.form['client_name'],
            direction=request.form['direction'],
            item_name=request.form['item_name'],
            model=request.form['model'],
            company_name=request.form['company_name'],
            project_name=request.form['project_name'],
            collected_by=request.form['collected_by'],
            quantity=int(request.form['quantity'])
        )
        db.session.add(item)
        db.session.commit()
        flash('New stock item added!', 'success')
        return redirect(url_for('main.stock_report'))
    return render_template('add_stock.html')

@main.route('/stock_summary', methods=['GET', 'POST'])
@login_required
def stock_summary():
    selected_item = request.form.get('item_name')
    selected_model = request.form.get('model')

    base_query = db.session.query(
        StockItem.item_name,
        StockItem.model,
        func.sum(case((StockItem.direction == 'IN', StockItem.quantity), else_=0)).label('total_in'),
        func.sum(case((StockItem.direction == 'OUT', StockItem.quantity), else_=0)).label('total_out'),
        (func.sum(case((StockItem.direction == 'IN', StockItem.quantity), else_=0)) -
         func.sum(case((StockItem.direction == 'OUT', StockItem.quantity), else_=0))).label('closing_balance')
    ).group_by(StockItem.item_name, StockItem.model)

    if selected_item:
        base_query = base_query.filter(StockItem.item_name == selected_item)
    if selected_model:
        base_query = base_query.filter(StockItem.model == selected_model)

    summary_data = base_query.all()
    item_names = db.session.query(StockItem.item_name).distinct().all()
    models = db.session.query(StockItem.model).distinct().all()
    total_in = sum(row.total_in or 0 for row in summary_data)
    total_out = sum(row.total_out or 0 for row in summary_data)
    closing_balance = total_in - total_out

    detailed_items = []
    if selected_item and selected_model:
        detailed_items = StockItem.query.filter_by(
            item_name=selected_item,
            model=selected_model
        ).order_by(StockItem.timestamp.asc()).all()

    return render_template(
        'stock_summary.html',
        summary_data=summary_data,
        item_names=[item[0] for item in item_names],
        models=[model[0] for model in models],
        selected_item=selected_item,
        selected_model=selected_model,
        total_in=total_in,
        total_out=total_out,
        closing_balance=closing_balance,
        detailed_items=detailed_items
    )

@main.route('/delete_stock/<int:stock_id>', methods=['POST'])
@login_required
def delete_stock(stock_id):
    stock = StockItem.query.get_or_404(stock_id)
    db.session.delete(stock)
    db.session.commit()
    flash('Stock entry deleted successfully.', 'success')
    return redirect(url_for('main.stock_report'))

@main.route('/download_sample_excel')
@login_required
def download_sample_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample Stock"
    ws.append(['Item Name', 'Company Name', 'Model', 'Client', 'Project', 'Quantity', 'Direction'])
    ws.append(['FMC', 'TELTONIKA', '920', 'Client A', 'Project Alpha', '10', 'IN'])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        download_name='sample_stock_upload.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
